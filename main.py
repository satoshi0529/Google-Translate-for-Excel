from googletrans import Translator
import openpyxl as excel

#Excelファイルの指定
test_file = excel.load_workbook("C:\\Users/y_sat/Desktop/test.xlsx",data_only=True)
#Sheetの選択
sheet = test_file.worksheets[0]
#excelの最終行を取得
max_row = sheet.max_row

translator = Translator()

for i in range(1,max_row+1):
    bf_trans = sheet.cell(row=i,column=1).value
    trans_text = translator.translate(bf_trans,dest='ja')
    sheet.cell(row=i,column=2).value = str(trans_text.text)

#excelファイルを閉じて終了
test_file.save("C:\\Users/y_sat/Desktop/test.xlsx")


